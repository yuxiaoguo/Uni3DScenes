"""
# Copyright (c) Microsoft Corporation.
# Licensed under the MIT License.
"""
import os
from typing import Dict

import openpyxl


class DistributionTable:
    """
    The IO interface for distribution table
    """
    def __init__(self, excel_path: str) -> None:
        self.excel_path = excel_path

        dir_path = os.path.dirname(self.excel_path)
        os.makedirs(dir_path, exist_ok=True)

        self.work_book = openpyxl.Workbook()
        self.work_book.remove(self.work_book.active)

    def write_overall(self, label_set, overall, sheet_name='Overall'):
        """
        Write overall information to a excel file
        """

        if overall is not None:
            sheet_oa = self.work_book.create_sheet(sheet_name)
            for c_idx, label in enumerate(label_set):
                sheet_oa.cell(c_idx + 1, 1, label)
            for c_idx, dist in enumerate(overall):
                sheet_oa.cell(c_idx + 1, 2, dist)

    def write_items(self, label_set, items, sheet_name='Items'):
        """
        Write items to a excel file
        """
        if items is not None:
            sheet_items = self.work_book.create_sheet(sheet_name)
            for c_idx, label in enumerate(label_set):
                sheet_items.cell(1, c_idx + 2, label)
            for r_idx, (r_key, r_value) in enumerate(items.items()):
                sheet_items.cell(r_idx + 2, 1, r_key)
                for c_idx, c_v in enumerate(r_value):
                    sheet_items.cell(r_idx + 2, c_idx + 2, c_v)

    def close(self):
        """
        Close and save the table
        """
        self.work_book.save(self.excel_path)

    def read(self):
        """
        Read from a excel file
        """
